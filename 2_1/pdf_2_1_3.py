import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
from jinja2 import Environment, FileSystemLoader
import pathlib
import pdfkit
import matplotlib.pyplot as plt
import numpy as np


class Vacancy:
    """
    Класс для вакансии и ее характеристик.
    Attributes:
        name (str): Название вакансии
        salary_from (int): Нижняя граница вилки оклада
        salary_to (int): Верхняя граница вилки оклада
        salary_currency (str): Валюта оклада
        salary_average (float): Средняя зарплата
        area_name (str): Месторасположение вакансии
        year (int): Год публикации
        currency_to_rub (dict): Словарь с курсом валют
    """
    def __init__(self, vacancy):
        """
        Инициализирует объект Vacancy.
        Args:
            vacancy (dict): словарь с характеристиками вакансии
        """
        self.name = vacancy['name']
        self.salary_from = int(float(vacancy['salary_from']))
        self.salary_to = int(float(vacancy['salary_to']))
        self.salary_currency = vacancy['salary_currency']
        self.salary_average = self.currency_to_rub[self.salary_currency] * (self.salary_from + self.salary_to) / 2
        self.area_name = vacancy['area_name']
        self.year = int(vacancy['published_at'][:4])

    currency_to_rub = {
        "AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76,
        "KZT": 0.13, "RUR": 1, "UAH": 1.64, "USD": 60.66, "UZS": 0.0055,
    }


class DataSet:
    """
    Считывает и обрабатывает данные из csv файла.
    Attributes:
        file_name (str): Название файла
        vacancy_name (str): Название вакансии
    """
    def __init__(self, file_name, vacancy_name):
        """
        Инициализирует объект DataSet.
        Args:
            file_name (str): Название файла
            vacancy_name (str): Название вакансии
        """
        self.file_name = file_name
        self.vacancy_name = vacancy_name

    def csv_reader(self):
        """
        Считывает csv файл и записывает данные в словарь.
        Returns:
            dict: Данные из csv файла
        """
        file = csv.reader(open(self.file_name, encoding='utf_8_sig'))
        title = next(file)
        title_length = len(title)
        for row in file:
            if len(row) == title_length and '' not in row:
                yield dict(zip(title, row))

    @staticmethod
    def average_for_salary(dict):
        """
        Находит среднюю зарплату в словаре.
        Args:
            dict (dict): словарь со средними зарплатами
        Returns:
             dict: словарь со средними зарплатами
        """
        answer = {}
        for key, values in dict.items():
            answer[key] = int(sum(values) / len(values))
        return answer

    @staticmethod
    def augmentation(vacancy, amount, dict):
        """
        Добавляет данные к словарю
        Args:
            vacancy (int): год вакансии
            amount (list): средняя зарплата
            dict (dict): словарь с зарплатами
        """
        if vacancy in dict:
            dict[vacancy] += amount
        else:
            dict[vacancy] = amount

    def get_statistic(self):
        """
        Формирует данные для статтистики.
        Returns:
            tuple: Солержит в себе словари с данными статистик
        """
        salary, salary_of_vacancy_name, salary_city, vacancies_count = {}, {}, {}, 0

        for vacancy_dict in self.csv_reader():
            vacancy = Vacancy(vacancy_dict)
            self.augmentation(vacancy.year, [vacancy.salary_average], salary)
            if vacancy.name.find(self.vacancy_name) != -1:
                self.augmentation(vacancy.year, [vacancy.salary_average], salary_of_vacancy_name)
            self.augmentation(vacancy.area_name, [vacancy.salary_average], salary_city)
            vacancies_count += 1

        number_vacancies = dict([(key, len(value)) for key, value in salary.items()])
        number_vacancies_by_name = dict([(key, len(value)) for key, value in salary_of_vacancy_name.items()])

        if not salary_of_vacancy_name:
            salary_of_vacancy_name = dict([(key, [0]) for key, value in salary.items()])
            number_vacancies_by_name = dict([(key, 0) for key, value in number_vacancies.items()])

        value = self.average_for_salary(salary)
        value2 = self.average_for_salary(salary_of_vacancy_name)
        value3 = self.average_for_salary(salary_city)
        value4 = {}
        for year, salaries in salary_city.items():
            value4[year] = round(len(salaries) / vacancies_count, 4)
        value4 = list(filter(lambda a: a[-1] >= 0.01, [(key, value) for key, value in value4.items()]))
        value4.sort(key=lambda a: a[-1], reverse=True)
        value5 = value4.copy()
        value5 = dict(value5[:10])
        value4 = dict(value4)
        value3 = list(filter(lambda a: a[0] in list(value4.keys()), [(key, value) for key, value in value3.items()]))
        value3.sort(key=lambda a: a[-1], reverse=True)
        value3 = dict(value3[:10])

        return value, number_vacancies, value2, number_vacancies_by_name, value3, value5

    @staticmethod
    def print_statistic(value1, value2, value3, value4, value5, value6):
        """
        Выводит данные в консоль.
        """
        print('Динамика уровня зарплат по годам: {0}'.format(value1))
        print('Динамика количества вакансий по годам: {0}'.format(value2))
        print('Динамика уровня зарплат по годам для выбранной профессии: {0}'.format(value3))
        print('Динамика количества вакансий по годам для выбранной профессии: {0}'.format(value4))
        print('Уровень зарплат по городам (в порядке убывания): {0}'.format(value5))
        print('Доля вакансий по городам (в порядке убывания): {0}'.format(value6))


class Report:
    """
    Выводит данные в excel, png, pdf , pdfkit.
    Attributes:
        workbook (Workbook): Инициализатор книги Excel
        vacancy_name (str): Название вакансии
        value1 (dict): Зарплаты по годам
        value2 (dict): Количество вакансий по годам
        value3 (dict): Зарплаты по годам для выбранной профессии
        value4 (dict): Количество вакансий по годам для выбранной профессии
        value5 (dict): Зарплаты по городам
        value6 (dict): Доли вакансий по городам
    """
    def __init__(self, vacancy_name, value1, value2, value3, value4, value5, value6):
        """
        Иницализация класса для вывода данных.
        Args:
            vacancy_name (str): Название вакансии
            value1 (dict): Зарплаты по годам
            value2 (dict): Количество вакансий по годам
            value3 (dict): Зарплаты по годам для выбранной профессии
            value4 (dict): Количество вакансий по годам для выбранной профессии
            value5 (dict): Зарплаты по городам
            value6 (dict): Доли вакансий по городам
        """
        self.workbook = Workbook()
        self.vacancy_name = vacancy_name
        self.value1 = value1
        self.value2 = value2
        self.value3 = value3
        self.value4 = value4
        self.value5 = value5
        self.value6 = value6

    def generate_excel(self):
        """
        Генерирует таблички в excel файле на двух страницах используя библиотеку openpyxl.
        Creates:
            Workbook: Файл report.xlsx с двумя страницами и таблицами в них
        """
        page1 = self.workbook.active
        page1.title = 'Статистика по годам'
        page1.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancy_name,
                      'Количество вакансий', 'Количество вакансий - ' + self.vacancy_name])
        for year in self.value1.keys():
            page1.append([year, self.value1[year], self.value3[year], self.value2[year], self.value4[year]])

        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancy_name,
                 ' Количество вакансий', ' Количество вакансий - ' + self.vacancy_name]]
        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                if len(column_widths) <= i:
                    column_widths += [len(cell)]
                else:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)

        for i, column_width in enumerate(column_widths, 1):
            page1.column_dimensions[get_column_letter(i)].width = column_width + 2

        page2 = self.workbook.create_sheet('Статистика по городам')
        data = [['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']]
        for (city1, value1), (city2, value2) in zip(self.value5.items(), self.value6.items()):
            data.append([city1, value1, '', city2, value2])
        for row in data:
            page2.append(row)
        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                cell = str(cell)
                if len(column_widths) <= i:
                    column_widths += [len(cell)]
                else:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)

        for i, column_width in enumerate(column_widths, 1):
            page2.column_dimensions[get_column_letter(i)].width = column_width + 2

        font_bold = Font(bold=True)
        for col in 'ABCDE':
            page1[col + '1'].font = font_bold
            page2[col + '1'].font = font_bold

        for index, _ in enumerate(self.value5):
            page2['E' + str(index + 2)].number_format = '0.00%'

        thin = Side(border_style='thin', color='00000000')

        for row in range(len(data)):
            for col in 'ABDE':
                page2[col + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

        for row, _ in enumerate(self.value1):
            for col in 'ABCDE':
                page1[col + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

        self.workbook.save(filename='report.xlsx')

    def generate_image(self):
        """
        Генерирует изображение с 4 графиками используя библиотеки matplotlib, numpy.
        Creates:
            png: Файл graph.png с 4 графиками
        """
        fig, ((graphic1, graphic2), (graphic3, graphic4)) = plt.subplots(nrows=2, ncols=2)

        graphic1.set_title('Уровень зарплат по годам', fontdict={'fontsize': 8})
        bar1 = graphic1.bar(np.array(list(self.value1.keys())) - 0.4, self.value1.values(), width=0.4)
        bar2 = graphic1.bar(np.array(list(self.value1.keys())), self.value3.values(), width=0.4)
        graphic1.legend((bar1[0], bar2[0]), ('средняя з/п', 'з/п ' + self.vacancy_name.lower()), prop={'size': 8})
        graphic1.set_xticks(np.array(list(self.value1.keys())) - 0.2, list(self.value1.keys()), rotation=90)
        graphic1.grid(axis='y')
        graphic1.xaxis.set_tick_params(labelsize=8)
        graphic1.yaxis.set_tick_params(labelsize=8)

        graphic2.set_title('Количество вакансий по годам', fontdict={'fontsize': 8})
        bar1 = graphic2.bar(np.array(list(self.value2.keys())) - 0.4, self.value2.values(), width=0.4)
        bar2 = graphic2.bar(np.array(list(self.value2.keys())), self.value4.values(), width=0.4)
        graphic2.legend((bar1[0], bar2[0]), ('Количество вакансий', 'Количество вакансий\n' +
                                             self.vacancy_name.lower()), prop={'size': 8})
        graphic2.set_xticks(np.array(list(self.value2.keys())) - 0.2, list(self.value2.keys()), rotation=90)
        graphic2.grid(axis='y')
        graphic2.xaxis.set_tick_params(labelsize=8)
        graphic2.yaxis.set_tick_params(labelsize=8)

        graphic3.set_title('Уровень зарплат по городам', fontdict={'fontsize': 8})
        bar1 = list([str(a).replace(' ', '\n').replace('-', '-\n') for a in reversed(list(self.value5.keys()))])
        bar2 = list(reversed(list(self.value5.values())))
        graphic3.barh(bar1, bar2, color='blue', height=0.5, align='center')
        graphic3.yaxis.set_tick_params(labelsize=6)
        graphic3.xaxis.set_tick_params(labelsize=8)
        graphic3.grid(axis='x')

        graphic4.set_title('Доля вакансий по городам', fontdict={'fontsize': 8})
        other = 1 - sum([value for value in self.value6.values()])
        self.value6 = dict(sorted(self.value6.items(), key=lambda x: x[1]))
        graphic4.pie(list(self.value6.values()) + [other], labels=list(self.value6.keys()) + ['Другие'],
                     textprops={'fontsize': 6})
        plt.tight_layout()
        plt.savefig('graph.png')

    def generate_pdf(self):
        """
        Генерирует pdf файл с 4 графиками и 2 таблицами, при помощи библиотек pdfkit, jinja2, pathlib и шаблона в формате html.
        Creates:
            pdf: Файл report.pdf с 4 графиками и 2 таблицами
        """
        environment = Environment(loader=FileSystemLoader('..'))
        template = environment.get_template("pdf_template.html")
        values = []
        for key in self.value6:
            self.value6[key] = round(self.value6[key] * 100, 2)
        for year in self.value1.keys():
            values.append([year, self.value1[year], self.value2[year], self.value3[year], self.value4[year]])
        pdf_template = template.render({'name': self.vacancy_name,
                                        'path': '{0}/{1}'.format(pathlib.Path(__file__).parent.resolve(),
                                                                 'graph.png'), 'values': values,
                                        'value5': self.value5, 'value6': self.value6})
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})


class InputConnect:
    """
    Обрабатывает данные введенные пользователем.
    Attributes:
        file_name (str): Название файла
        vacancy_name (str): Название вакансии
    """
    def __init__(self):
        """
        Инициализирует класс, выводит статистику в консоль, а так же запускает создание xlsx, png, pdf файлов.
        """
        self.file_name = input('Введите название файла: ')
        self.vacancy_name = input('Введите название профессии: ')

        dataset = DataSet(self.file_name, self.vacancy_name)
        value1, value2, value3, value4, value5, value6 = dataset.get_statistic()
        dataset.print_statistic(value1, value2, value3, value4, value5, value6)

        report = Report(self.vacancy_name, value1, value2, value3, value4, value5, value6)
        report.generate_excel()
        report.generate_image()
        report.generate_pdf()


def get_pdf():
    """
    Запускает программу.
    """
    InputConnect()
