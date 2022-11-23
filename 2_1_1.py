import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side


class Vacancy:
    def __init__(self, vacancy):
        self.name = vacancy['name']
        self.salary_from = int(float(vacancy['salary_from']))
        self.salary_to = int(float(vacancy['salary_to']))
        self.salary_currency = vacancy['salary_currency']
        self.salary_average = self.currency_to_rub[self.salary_currency] * (self.salary_from + self.salary_to) / 2
        self.area_name = vacancy['area_name']
        self.year = int(vacancy['published_at'][:4])

    currency_to_rub = {
        "AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76,
        "KZT": 0.13, "RUR": 1, "UAH": 1.64, "USD": 60.66, "UZS": 0.0055,
    }


class DataSet:
    def __init__(self, file_name, vacancy_name):
        self.file_name = file_name
        self.vacancy_name = vacancy_name

    def csv_reader(self):
        file = csv.reader(open(self.file_name, encoding='utf_8_sig'))
        title = next(file)
        title_length = len(title)
        for row in file:
            if len(row) == title_length and '' not in row:
                yield dict(zip(title, row))

    @staticmethod
    def average_for_salary(dict):
        answer = {}
        for key, values in dict.items():
            answer[key] = int(sum(values) / len(values))
        return answer

    @staticmethod
    def augmentation(vacancy, amount, dict):
        if vacancy in dict:
            dict[vacancy] += amount
        else:
            dict[vacancy] = amount

    def get_statistic(self):
        salary, salary_of_vacancy_name, salary_city, vacancies_count = {}, {}, {}, 0

        for vacancy_dict in self.csv_reader():
            vacancy = Vacancy(vacancy_dict)
            self.augmentation(vacancy.year, [vacancy.salary_average], salary)
            if vacancy.name.find(self.vacancy_name) != -1:
                self.augmentation(vacancy.year, [vacancy.salary_average], salary_of_vacancy_name)
            self.augmentation(vacancy.area_name, [vacancy.salary_average], salary_city)
            vacancies_count += 1

        number_vacancies = dict([(key, len(value)) for key, value in salary.items()])
        number_vacancies_by_name = dict([(key, len(value)) for key, value in salary_of_vacancy_name.items()])

        if not salary_of_vacancy_name:
            salary_of_vacancy_name = dict([(key, [0]) for key, value in salary.items()])
            number_vacancies_by_name = dict([(key, 0) for key, value in number_vacancies.items()])

        value = self.average_for_salary(salary)
        value2 = self.average_for_salary(salary_of_vacancy_name)
        value3 = self.average_for_salary(salary_city)
        value4 = {}
        for year, salaries in salary_city.items():
            value4[year] = round(len(salaries) / vacancies_count, 4)
        value4 = list(filter(lambda a: a[-1] >= 0.01, [(key, value) for key, value in value4.items()]))
        value4.sort(key=lambda a: a[-1], reverse=True)
        value5 = value4.copy()
        value5 = dict(value5[:10])
        value4 = dict(value4)
        value3 = list(filter(lambda a: a[0] in list(value4.keys()), [(key, value) for key, value in value3.items()]))
        value3.sort(key=lambda a: a[-1], reverse=True)
        value3 = dict(value3[:10])

        return value, number_vacancies, value2, number_vacancies_by_name, value3, value5

    @staticmethod
    def print_statistic(value1, value2, value3, value4, value5, value6):
        print('Динамика уровня зарплат по годам: {0}'.format(value1))
        print('Динамика количества вакансий по годам: {0}'.format(value2))
        print('Динамика уровня зарплат по годам для выбранной профессии: {0}'.format(value3))
        print('Динамика количества вакансий по годам для выбранной профессии: {0}'.format(value4))
        print('Уровень зарплат по городам (в порядке убывания): {0}'.format(value5))
        print('Доля вакансий по городам (в порядке убывания): {0}'.format(value6))


class Report:
    def __init__(self, vacancy_name, value1, value2, value3, value4, value5, value6):
        self.workbook = Workbook()
        self.vacancy_name = vacancy_name
        self.value1 = value1
        self.value2 = value2
        self.value3 = value3
        self.value4 = value4
        self.value5 = value5
        self.value6 = value6

    def generate_excel(self):
        page1 = self.workbook.active
        page1.title = 'Статистика по годам'
        page1.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancy_name,
                    'Количество вакансий', 'Количество вакансий - ' + self.vacancy_name])
        for year in self.value1.keys():
            page1.append([year, self.value1[year], self.value3[year], self.value2[year], self.value4[year]])

        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancy_name,
                 ' Количество вакансий', ' Количество вакансий - ' + self.vacancy_name]]
        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                if len(column_widths) <= i:
                    column_widths += [len(cell)]
                else:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)

        for i, column_width in enumerate(column_widths, 1):
            page1.column_dimensions[get_column_letter(i)].width = column_width + 2

        page2 = self.workbook.create_sheet('Статистика по городам')
        data = [['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']]
        for (city1, value1), (city2, value2) in zip(self.value5.items(), self.value6.items()):
            data.append([city1, value1, '', city2, value2])
        for row in data:
            page2.append(row)
        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                cell = str(cell)
                if len(column_widths) <= i:
                    column_widths += [len(cell)]
                else:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)

        for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
            page2.column_dimensions[get_column_letter(i)].width = column_width + 2

        font_bold = Font(bold=True)
        for col in 'ABCDE':
            page1[col + '1'].font = font_bold
            page2[col + '1'].font = font_bold

        for index, _ in enumerate(self.value5):
            page2['E' + str(index + 2)].number_format = '0.00%'

        thin = Side(border_style='thin', color='00000000')

        for row in range(len(data)):
            for col in 'ABDE':
                page2[col + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

        self.value1[1] = 1
        for row, _ in enumerate(self.value1):
            for col in 'ABCDE':
                page1[col + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

        self.workbook.save('report.xlsx')


class InputConnect:
    def __init__(self):
        self.file_name = input('Введите название файла: ')
        self.vacancy_name = input('Введите название профессии: ')

        dataset = DataSet(self.file_name, self.vacancy_name)
        value1, value2, value3, value4, value5, value6 = dataset.get_statistic()
        dataset.print_statistic(value1, value2, value3, value4, value5, value6)

        report = Report(self.vacancy_name, value1, value2, value3, value4, value5, value6)
        report.generate_excel()


if __name__ == '__main__':
    InputConnect()